"""
证件检查工具模块
从现有系统复用核心功能
支持简单格式（每行一个证件）和复杂格式（每行一个人员，多证件列）
"""
from datetime import datetime, timedelta
from typing import Optional, Tuple, Dict, List
import pandas as pd

# Import STATUS_MAP from config to avoid duplication
try:
    from config import STATUS_MAP, URGENT_DAYS, WARNING_DAYS
except ImportError:
    # Fallback values if config is not available
    STATUS_MAP = {
        'expired': {'label': '已过期', 'icon': 'X', 'color': '#C0392B'},
        'urgent': {'label': '紧急', 'icon': '!', 'color': '#E74C3C'},
        'warning': {'label': '预警', 'icon': '?', 'color': '#F39C12'},
        'normal': {'label': '正常', 'icon': 'OK', 'color': '#27AE60'},
        'unknown': {'label': '日期无效', 'icon': '?', 'color': '#95A5A6'}
    }
    URGENT_DAYS = 30
    WARNING_DAYS = 90


def calculate_days_remaining(expiry_date, query_date=None):
    """
    计算剩余天数

    Args:
        expiry_date: 到期日期 (datetime、Timestamp或字符串)
        query_date: 查询日期，默认为今天

    Returns:
        剩余天数（负数表示已过期）
    """
    if query_date is None:
        query_date = datetime.now().date()

    # 处理pandas Timestamp类型
    try:
        import pandas as pd
        if isinstance(expiry_date, pd.Timestamp):
            expiry_date = expiry_date.to_pydatetime()
        elif pd.isna(expiry_date):
            return None
    except ImportError:
        pass

    # 处理整数类型（可能是Excel日期序列号）
    if isinstance(expiry_date, (int, float)) and expiry_date > 0:
        try:
            import pandas as pd
            expiry_date = pd.to_datetime('1900-01-01') + pd.to_timedelta(expiry_date - 2, unit='D')
        except:
            return None

    if isinstance(expiry_date, str):
        try:
            expiry_date = datetime.strptime(expiry_date, "%Y-%m-%d")
        except ValueError:
            try:
                expiry_date = datetime.strptime(expiry_date, "%Y/%m/%d")
            except ValueError:
                return None

    # 统一转换为date对象进行计算
    if isinstance(expiry_date, datetime):
        expiry_date = expiry_date.date()

    if isinstance(query_date, datetime):
        query_date = query_date.date()

    return (expiry_date - query_date).days


def get_status_indicator(days_remaining: Optional[int]) -> Dict:
    """
    根据剩余天数返回状态信息
    使用STATUS_MAP from config.py避免重复定义

    Args:
        days_remaining: 剩余天数

    Returns:
        包含label, icon, color, status的字典
    """
    if days_remaining is None:
        return {**STATUS_MAP['unknown'], 'status': 'unknown'}
    elif days_remaining < 0:
        return {**STATUS_MAP['expired'], 'status': 'expired'}
    elif days_remaining <= URGENT_DAYS:
        return {**STATUS_MAP['urgent'], 'status': 'urgent'}
    elif days_remaining <= WARNING_DAYS:
        return {**STATUS_MAP['warning'], 'status': 'warning'}
    else:
        return {**STATUS_MAP['normal'], 'status': 'normal'}


def detect_excel_format(file_path) -> str:
    """
    自动检测Excel文件格式

    Args:
        file_path: Excel文件路径

    Returns:
        'simple' - 简单格式（每行一个证件，有表头）
        'complex' - 复杂格式（每行一个人员，多证件列，无标准表头）
    """
    df = pd.read_excel(file_path, sheet_name=0, header=None, engine='openpyxl')

    # 检查第一行是否包含标题特征
    first_row = df.iloc[0].astype(str).tolist()

    # 复杂格式特征：第一行包含"人员证件台帐"或"EBS"
    if any('人员证件' in str(cell) or 'EBS' in str(cell) for cell in first_row):
        return 'complex'

    # 简单格式特征：第一列包含"姓名"或类似标准表头
    try:
        df_with_header = pd.read_excel(file_path, sheet_name=0, header=0)
        if '姓名' in df_with_header.columns and '证件名称' in df_with_header.columns:
            return 'simple'
    except:
        pass

    return 'simple'


def get_sheet_names(file_path) -> list:
    """获取Excel文件中所有sheet名称"""
    xl = pd.ExcelFile(file_path)
    return xl.sheet_names


def parse_simple_format(file_path: str, sheet_name: str = 0) -> Tuple[List[Dict], Dict]:
    """
    解析简单格式Excel文件（每行一个证件）

    Args:
        file_path: Excel文件路径
        sheet_name: 指定要解析的sheet索引

    Returns:
        (certificates, statistics) - 证件列表和统计信息
    """
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    required_columns = ["姓名", "部门", "证件名称", "证件号码", "到期日期", "邮箱", "手机号"]

    # 验证列名
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Excel缺少必需列: {', '.join(missing_columns)}")

    certificates = []
    total_count = len(df)
    stats = {'total': 0, 'expired': 0, 'urgent': 0, 'warning': 0, 'normal': 0}

    # Check optional columns once before loop (efficiency optimization)
    has_position_column = '岗位' in df.columns
    has_issue_date_column = '取证日期' in df.columns or '发证日期' in df.columns

    # 确定发证日期列名
    issue_date_col = None
    if '发证日期' in df.columns:
        issue_date_col = '发证日期'
    elif '取证日期' in df.columns:
        issue_date_col = '取证日期'

    for _, row in df.iterrows():
        expiry_date = row.get('到期日期')
        days_remaining = calculate_days_remaining(expiry_date)

        status_info = get_status_indicator(days_remaining)

        cert = {
            'name': str(row.get('姓名', '')).strip(),
            'department': str(row.get('部门', '')).strip(),
            'position': str(row.get('岗位', '')).strip() if has_position_column else '',
            'certificate_name': str(row.get('证件名称', '')).strip(),
            'certificate_number': str(row.get('证件号码', '')).strip(),
            'expiry_date': _format_date(expiry_date),
            'issue_date': _format_date(row.get(issue_date_col)) if has_issue_date_column else '',
            'email': str(row.get('邮箱', '')).strip(),
            'phone': str(row.get('手机号', '')).strip(),
            'days_remaining': days_remaining,
            'status': status_info['status'],
            'status_label': status_info['label'],
            'status_icon': status_info['icon'],
            'status_color': status_info['color']
        }

        if days_remaining is not None:
            certificates.append(cert)
            stats['total'] += 1
            stats[status_info['status']] += 1

    return certificates, stats


def parse_complex_format(file_path: str, sheet_name: int = 0) -> Tuple[List[Dict], Dict]:
    """
    解析复杂格式Excel文件（每行一个人员，多证件列）

    Args:
        file_path: Excel文件路径
        sheet_name: 指定要解析的sheet索引

    Returns:
        (certificates, statistics) - 证件列表和统计信息
    """
    # 证件类型列定义
    CERTIFICATE_TYPES = [
        {'name': 'IADC/IWCF', 'num_col': 9, 'issue_col': 10, 'exp_col': 11},
        {'name': 'HSE证(H2S)', 'num_col': 12, 'issue_col': 13, 'exp_col': 14},
        {'name': '急救证', 'num_col': 15, 'issue_col': 16, 'exp_col': 17},
        {'name': '消防证', 'num_col': 18, 'issue_col': 19, 'exp_col': 20},
        {'name': '司索指挥证', 'num_col': 21, 'issue_col': 22, 'exp_col': 23},
        {'name': '防恐证', 'num_col': 24, 'issue_col': 25, 'exp_col': 26},
        {'name': '健康证', 'num_col': 30, 'issue_col': 31, 'exp_col': 32},
        {'name': '岗位证', 'num_col': 33, 'issue_col': 34, 'exp_col': 35},
    ]

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
    certificates = []
    stats = {'total': 0, 'expired': 0, 'urgent': 0, 'warning': 0, 'normal': 0}

    # 数据从第4行开始（pandas索引3，因为前3行是标题）
    # 使用range(3, ...)确保从正确的数据行开始
    for idx in range(3, len(df)):
        row = df.iloc[idx]
        name = str(row[2]).strip() if pd.notna(row[2]) else ''

        if not name or name == 'nan':
            continue

        unit = str(row[1]).strip() if pd.notna(row[1]) else ''
        phone = str(row[5]).strip() if pd.notna(row[5]) else ''

        for cert in CERTIFICATE_TYPES:
            cert_num = str(row[cert['num_col']]).strip() if pd.notna(row[cert['num_col']]) else ''
            exp_date = row[cert['exp_col']]
            issue_date = row[cert['issue_col']]  # 读取发证日期

            if not cert_num or cert_num == 'nan':
                continue

            days_remaining = calculate_days_remaining(exp_date)

            if days_remaining is None:
                continue

            status_info = get_status_indicator(days_remaining)

            cert = {
                'name': name,
                'department': unit,
                'position': '',  # 复杂格式无岗位列
                'certificate_name': cert['name'],
                'certificate_number': cert_num,
                'expiry_date': _format_date(exp_date),
                'issue_date': _format_date(issue_date),  # 格式化发证日期
                'email': '',
                'phone': phone,
                'days_remaining': days_remaining,
                'status': status_info['status'],
                'status_label': status_info['label'],
                'status_icon': status_info['icon'],
                'status_color': status_info['color']
            }

            certificates.append(cert)
            stats['total'] += 1
            stats[status_info['status']] += 1

    return certificates, stats


def _format_date(date_val) -> str:
    """格式化日期为 YYYY-MM-DD 字符串"""
    if date_val is None:
        return ''
    if pd.isna(date_val):
        return ''

    if isinstance(date_val, str):
        return date_val

    try:
        if isinstance(date_val, pd.Timestamp):
            return date_val.strftime('%Y-%m-%d')
        if isinstance(date_val, datetime):
            return date_val.strftime('%Y-%m-%d')
    except:
        pass

    return str(date_val)


def parse_excel_file(file_path: str, sheet_name: str = None) -> Tuple[List[Dict], Dict]:
    """
    自动检测并解析Excel文件

    Args:
        file_path: Excel文件路径
        sheet_name: 指定要解析的sheet名称/索引

    Returns:
        (certificates, statistics) - 证件列表和统计信息
    """
    format_type = detect_excel_format(file_path)

    if sheet_name is None:
        sheet_name = 0

    if format_type == 'complex':
        return parse_complex_format(file_path, sheet_name)
    else:
        return parse_simple_format(file_path, sheet_name)


# ============ JSON Persistence ============

import json
import os
from typing import List


def save_to_json(data: List[Dict], filepath: str, metadata: Dict = None) -> bool:
    """保存证件数据到JSON文件"""
    try:
        dir_path = os.path.dirname(filepath)
        if dir_path:  # 只在有目录路径时才创建
            os.makedirs(dir_path, exist_ok=True)
        output = {'data': data}
        if metadata:
            output['metadata'] = metadata
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"Error saving to JSON: {e}")
        return False


def load_from_json(filepath: str) -> List[Dict]:
    """从JSON文件加载证件数据"""
    if not os.path.exists(filepath):
        return []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            result = json.load(f)
            # 兼容新旧格式
            if isinstance(result, dict):
                return result.get('data', [])
            return result
    except Exception as e:
        print(f"Error loading from JSON: {e}")
        return []


# ============ Excel Export ============

def export_to_excel(data: List[Dict], filepath: str) -> bool:
    """导出证件数据到Excel文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "证件数据"

        # 定义表头（使用发证日期作为列名）
        headers = ['姓名', '部门', '岗位', '证件名称', '证件号码', '发证日期', '到期日期', '剩余天数', '状态', '邮箱', '手机号']

        # 表头样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style

        # 写入数据
        for row_idx, cert in enumerate(data, start=2):
            ws.cell(row=row_idx, column=1, value=cert.get('name', ''))
            ws.cell(row=row_idx, column=2, value=cert.get('department', ''))
            ws.cell(row=row_idx, column=3, value=cert.get('position', ''))
            ws.cell(row=row_idx, column=4, value=cert.get('certificate_name', ''))
            ws.cell(row=row_idx, column=5, value=cert.get('certificate_number', ''))
            ws.cell(row=row_idx, column=6, value=cert.get('issue_date', ''))
            ws.cell(row=row_idx, column=7, value=cert.get('expiry_date', ''))
            ws.cell(row=row_idx, column=8, value=cert.get('days_remaining', ''))
            ws.cell(row=row_idx, column=9, value=cert.get('status_label', ''))
            ws.cell(row=row_idx, column=10, value=cert.get('email', ''))
            ws.cell(row=row_idx, column=11, value=cert.get('phone', ''))

            # 根据状态设置行背景色和字体颜色
            status = cert.get('status', 'normal')
            fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            font_color = '000000'  # 默认黑色
            bold = False

            if status == 'expired':
                # 已过期 - 深红色背景，白色字体
                fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
                font_color = 'FFFFFF'
                bold = True
            elif status == 'urgent':
                # 紧急 - 红色背景，白色字体
                fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
                font_color = 'FFFFFF'
                bold = True
            elif status == 'warning':
                # 预警 - 淡黄色背景，黑色字体
                fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
                font_color = '000000'
                bold = True
            elif status == 'normal':
                fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                font_color = '000000'

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border_style
                cell.fill = fill
                cell.font = Font(color=font_color, bold=bold)

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # 确保目录存在
        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        wb.save(filepath)
        return True
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        return False


def export_updated_original(original_filepath: str, output_filepath: str, certificates: List[Dict], metadata: Dict) -> bool:
    """
    导出更新后的原表格（保留原始Excel格式，使用当前数据更新）
    支持添加新员工行到原表格
    保留原表格的条件格式

    Args:
        original_filepath: 原始上传的Excel文件路径
        output_filepath: 输出文件路径
        certificates: 当前证件数据列表
        metadata: 上传元数据（包含格式类型、sheet名称等）

    Returns:
        是否成功
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill

        format_type = metadata.get('format_type', 'simple')

        if format_type == 'complex':
            # 复杂格式：直接加载原文件并更新数据单元格（保留所有格式）
            # data_only=False 保留公式和格式
            wb = openpyxl.load_workbook(original_filepath, data_only=False)

            # 获取正确的sheet
            sheet_name = metadata.get('sheet_name')
            if sheet_name and isinstance(sheet_name, str) and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # 注意：使用 data_only=False 加载时，条件格式会自动保留
            # 不需要手动备份和恢复

            # 证件类型列定义（pandas是0索引，openpyxl是1索引，所以需要+1）
            cert_types = [
                {'name': 'IADC/IWCF', 'num_col': 10, 'issue_col': 11, 'exp_col': 12},
                {'name': 'HSE证(H2S)', 'num_col': 13, 'issue_col': 14, 'exp_col': 15},
                {'name': '急救证', 'num_col': 16, 'issue_col': 17, 'exp_col': 18},
                {'name': '消防证', 'num_col': 19, 'issue_col': 20, 'exp_col': 21},
                {'name': '司索指挥证', 'num_col': 22, 'issue_col': 23, 'exp_col': 24},
                {'name': '防恐证', 'num_col': 25, 'issue_col': 26, 'exp_col': 27},
                {'name': '健康证', 'num_col': 31, 'issue_col': 32, 'exp_col': 33},
                {'name': '岗位证', 'num_col': 34, 'issue_col': 35, 'exp_col': 36},
            ]

            # 创建数据映射：(姓名, 部门, 证件名称) -> 证件数据
            cert_map = {}
            for cert in certificates:
                key = (cert.get('name', ''), cert.get('department', ''), cert.get('certificate_name', ''))
                cert_map[key] = cert

            # 创建在职人员集合：记录当前数据中有哪些（姓名，部门）
            current_people = set()
            for cert in certificates:
                current_people.add((cert.get('name', ''), cert.get('department', '')))

            # 记录已处理的证件（姓名，部门）
            processed_people = set()

            # 读取原文件结构，找到每个人员行的位置
            # 数据从第4行开始（前3行是标题行）
            row_idx = 4  # Excel数据从第4行开始
            max_rows = 1000

            while row_idx <= max_rows:
                name_cell = ws.cell(row=row_idx, column=3)
                name = str(name_cell.value).strip() if name_cell.value else ''

                if not name or name == 'None' or name == 'nan':
                    # 如果是空行，检查是否还有更多数据
                    has_data = False
                    for col in range(1, 36):
                        cell = ws.cell(row=row_idx, column=col)
                        if cell.value and str(cell.value).strip() and str(cell.value).strip() != 'nan':
                            has_data = True
                            break
                    if not has_data:
                        break  # 没有更多数据，退出
                    row_idx += 1
                    continue  # 跳过此空行，处理下一行

                dept_cell = ws.cell(row=row_idx, column=2)
                dept = str(dept_cell.value).strip() if dept_cell.value else ''

                person_key = (name, dept)

                # 检查此人是否还在当前数据中
                person_exists = person_key in current_people

                if not person_exists:
                    # 此人已被删除：清空所有证件数据并标记为已删除
                    # 标记姓名为已删除
                    name_cell.value = f'{name} (已删除)'
                    # 标记姓名为红色
                    name_cell.font = openpyxl.styles.Font(color='FF0000', bold=True)

                    # 清空所有证件列的数据
                    for cert_type in cert_types:
                        # 清空证件号码
                        ws.cell(row=row_idx, column=cert_type['num_col']).value = ''
                        # 清空发证日期
                        ws.cell(row=row_idx, column=cert_type['issue_col']).value = ''
                        # 清空到期日期
                        ws.cell(row=row_idx, column=cert_type['exp_col']).value = ''
                else:
                    # 此人仍在：更新证件数据
                    # 标记此人员已处理
                    processed_people.add((name, dept))

                    # 更新此行的所有证件
                    for cert_type in cert_types:
                        cert_key = (name, dept, cert_type['name'])

                        if cert_key in cert_map:
                            cert = cert_map[cert_key]

                            # 更新证件号码（保留原有格式）
                            num_cell = ws.cell(row=row_idx, column=cert_type['num_col'])
                            original_num_style = num_cell.style if hasattr(num_cell, 'style') else None
                            num_cell.value = cert.get('certificate_number', '')
                            if original_num_style:
                                try:
                                    num_cell.style = original_num_style
                                except:
                                    pass

                            # 更新发证日期（保留原有格式）
                            issue_cell = ws.cell(row=row_idx, column=cert_type['issue_col'])
                            issue_date = cert.get('issue_date', '')
                            if issue_date:
                                try:
                                    from datetime import datetime
                                    if isinstance(issue_date, str):
                                        issue_date = datetime.strptime(issue_date, '%Y-%m-%d')
                                    # 保存原有格式
                                    original_format = issue_cell.number_format
                                    issue_cell.value = issue_date
                                    # 只有在原格式不是日期格式时才设置
                                    if not original_format or original_format == 'General':
                                        issue_cell.number_format = 'YYYY-MM-DD'
                                    else:
                                        issue_cell.number_format = original_format
                                except:
                                    issue_cell.value = issue_date
                            else:
                                issue_cell.value = ''

                            # 更新到期日期（保留原有格式）
                            exp_cell = ws.cell(row=row_idx, column=cert_type['exp_col'])
                            exp_date = cert.get('expiry_date', '')
                            if exp_date:
                                try:
                                    if isinstance(exp_date, str):
                                        exp_date = datetime.strptime(exp_date, '%Y-%m-%d')
                                    # 保存原有格式
                                    original_format = exp_cell.number_format
                                    exp_cell.value = exp_date
                                    # 只有在原格式不是日期格式时才设置
                                    if not original_format or original_format == 'General':
                                        exp_cell.number_format = 'YYYY-MM-DD'
                                    else:
                                        exp_cell.number_format = original_format
                                except:
                                    exp_cell.value = exp_date
                            else:
                                exp_cell.value = ''
                        else:
                            # 此证件不存在于当前数据中（该证件被删除），清空该证件的数据
                            ws.cell(row=row_idx, column=cert_type['num_col']).value = ''
                            ws.cell(row=row_idx, column=cert_type['issue_col']).value = ''
                            ws.cell(row=row_idx, column=cert_type['exp_col']).value = ''

                row_idx += 1

            # 第二步：添加新员工行（不在原表格中的人员）
            # 按姓名+部门分组
            cert_by_person = {}
            for cert in certificates:
                person_key = (cert.get('name', ''), cert.get('department', ''))
                if person_key not in cert_by_person:
                    cert_by_person[person_key] = []
                cert_by_person[person_key].append(cert)

            # 找出新员工并添加行
            new_row_idx = row_idx
            for person_key, cert_list in cert_by_person.items():
                if person_key not in processed_people:
                    name, dept = person_key

                    # 添加新行，复制格式
                    for col in range(1, 36):
                        src_cell = ws.cell(row=row_idx-1, column=col)
                        dst_cell = ws.cell(row=new_row_idx, column=col)
                        if src_cell.has_style:
                            try:
                                dst_cell.font = openpyxl.styles.Font(
                                    name=src_cell.font.name,
                                    bold=src_cell.font.bold,
                                    italic=src_cell.font.italic
                                )
                                if src_cell.fill:
                                    dst_cell.fill = openpyxl.styles.PatternFill(
                                        start_color=src_cell.fill.start_color
                                    )
                                if src_cell.border:
                                    dst_cell.border = src_cell.border
                                if src_cell.alignment:
                                    dst_cell.alignment = openpyxl.styles.Alignment(
                                        horizontal=src_cell.alignment.horizontal,
                                        vertical=src_cell.alignment.vertical
                                    )
                            except:
                                pass

                    # 填写基本信息
                    ws.cell(row=new_row_idx, column=2).value = dept
                    ws.cell(row=new_row_idx, column=3).value = name
                    ws.cell(row=new_row_idx, column=5).value = cert_list[0].get('phone', '') if cert_list else ''

                    # 填写所有证件数据
                    for cert in cert_list:
                        cert_type_name = cert.get('certificate_name', '')

                        # 找到对应的证件类型配置
                        for ct in cert_types:
                            if ct['name'] == cert_type_name:
                                # 填写证件号码（复制上方行的格式）
                                num_src_cell = ws.cell(row=row_idx-1, column=ct['num_col'])
                                num_dst_cell = ws.cell(row=new_row_idx, column=ct['num_col'])
                                num_dst_cell.value = cert.get('certificate_number', '')
                                if num_src_cell and hasattr(num_src_cell, 'style'):
                                    try:
                                        num_dst_cell.style = num_src_cell.style
                                        num_dst_cell.number_format = num_src_cell.number_format
                                    except:
                                        pass

                                # 填写发证日期（复制上方行的格式）
                                issue_date = cert.get('issue_date', '')
                                if issue_date:
                                    try:
                                        from datetime import datetime
                                        if isinstance(issue_date, str):
                                            issue_date = datetime.strptime(issue_date, '%Y-%m-%d')
                                        issue_src_cell = ws.cell(row=row_idx-1, column=ct['issue_col'])
                                        issue_dst_cell = ws.cell(row=new_row_idx, column=ct['issue_col'])
                                        issue_dst_cell.value = issue_date
                                        if issue_src_cell and hasattr(issue_src_cell, 'number_format'):
                                            try:
                                                issue_dst_cell.number_format = issue_src_cell.number_format
                                                issue_dst_cell.style = issue_src_cell.style
                                            except:
                                                issue_dst_cell.number_format = 'YYYY-MM-DD'
                                        else:
                                            issue_dst_cell.number_format = 'YYYY-MM-DD'
                                    except:
                                        ws.cell(row=new_row_idx, column=ct['issue_col']).value = issue_date

                                # 填写到期日期（复制上方行的格式）
                                exp_date = cert.get('expiry_date', '')
                                if exp_date:
                                    try:
                                        if isinstance(exp_date, str):
                                            exp_date = datetime.strptime(exp_date, '%Y-%m-%d')
                                        exp_src_cell = ws.cell(row=row_idx-1, column=ct['exp_col'])
                                        exp_dst_cell = ws.cell(row=new_row_idx, column=ct['exp_col'])
                                        exp_dst_cell.value = exp_date
                                        if exp_src_cell and hasattr(exp_src_cell, 'number_format'):
                                            try:
                                                exp_dst_cell.number_format = exp_src_cell.number_format
                                                exp_dst_cell.style = exp_src_cell.style
                                            except:
                                                exp_dst_cell.number_format = 'YYYY-MM-DD'
                                        else:
                                            exp_dst_cell.number_format = 'YYYY-MM-DD'
                                    except:
                                        ws.cell(row=new_row_idx, column=ct['exp_col']).value = exp_date
                                break

                    new_row_idx += 1

            # 保存工作簿（条件格式会自动保留，因为使用了 data_only=False）
            wb.save(output_filepath)
            return True

        else:
            # 简单格式：使用标准格式导出
            return export_to_excel(certificates, output_filepath)

    except Exception as e:
        print(f"Error exporting updated original: {e}")
        import traceback
        traceback.print_exc()
        # 如果复杂格式更新失败，使用标准导出
        return export_to_excel(certificates, output_filepath)


# ============ Search ============

def search_certificates(certificates: List[Dict], search_term: str = '', status_filter: str = '') -> List[Dict]:
    """
    搜索和过滤证件数据

    Args:
        certificates: 证件列表
        search_term: 搜索关键词（姓名、部门、岗位、证件名称、证件号码）
        status_filter: 状态过滤（expired/urgent/warning/normal）

    Returns:
        过滤后的证件列表
    """
    results = certificates

    if search_term:
        search_term = search_term.lower()
        results = [
            c for c in results
            if (search_term in c.get('name', '').lower() or
                search_term in c.get('department', '').lower() or
                search_term in c.get('position', '').lower() or
                search_term in c.get('certificate_name', '').lower() or
                search_term in c.get('certificate_number', '').lower())
        ]

    if status_filter:
        results = [c for c in results if c.get('status') == status_filter]

    return results
